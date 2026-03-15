"""
Genera mapas curriculares estructurados a partir del Excel oficial del plan 2021.

Salidas:
- data/mapa_curricular_2021ID_real.json
- data/mapa_curricular_2021ID_real_completo.json
- data/equivalencias_legacy_2021ID.json
"""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


PROYECTO_ROOT = Path(__file__).resolve().parent.parent
RUTA_EXCEL = PROYECTO_ROOT / "data" / "MAPA IDEAL IDEIO 2021 (rev 2025).xlsx"
RUTA_MAPA_LEGACY = PROYECTO_ROOT / "data" / "mapa_curricular_2021ID_completo.json"
RUTA_SALIDA_BASE = PROYECTO_ROOT / "data" / "mapa_curricular_2021ID_real.json"
RUTA_SALIDA_COMPLETA = PROYECTO_ROOT / "data" / "mapa_curricular_2021ID_real_completo.json"
RUTA_EQUIVALENCIAS = PROYECTO_ROOT / "data" / "equivalencias_legacy_2021ID.json"


FILAS_SEMESTRE = {
    8: 1,
    10: 2,
    12: 3,
    14: 4,
    16: 5,
    18: 6,
    20: 7,
    22: 8,
}

CODIGOS_CATEGORIA = {
    "B": "BASICA",
    "E": "ELECCION_LIBRE",
    "P": "PREESPECIALIDAD",
}

PATRON_MATERIA = re.compile(r"^([A-Z]{2,4}\d{4})\s*(.+)$", re.DOTALL)

# Equivalencias manuales seguras entre el mapa simplificado anterior y el plan real.
EQUIVALENCIAS_MANUALES = {
    "ID0001": "II0002",
    "ID0106": "DP0001",
    "ID0304": "IT3472",
    "ID0605": "ID3476",
    "ID0701": "IT0427",
}


def normalizar(texto: str) -> str:
    texto = str(texto or "").strip().lower()
    texto = "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    texto = re.sub(r"\([^)]*\)", "", texto)
    texto = re.sub(r"[^a-z0-9]+", " ", texto).strip()
    return texto


def limpiar_nombre(nombre_bruto: str) -> str:
    nombre = re.sub(r"\s+", " ", nombre_bruto.replace("\n", " ")).strip()
    return nombre


def leer_mapa_excel(ruta_excel: Path) -> Dict[str, Dict]:
    ws = load_workbook(ruta_excel, data_only=True).active
    mapa = {}

    for fila, semestre in FILAS_SEMESTRE.items():
        ciclo_anual = ((semestre - 1) // 2) + 1

        for columna in range(1, ws.max_column + 1):
            valor = ws.cell(fila, columna).value
            if not isinstance(valor, str):
                continue

            texto = valor.strip()
            match = PATRON_MATERIA.match(texto)
            if not match:
                continue

            clave = match.group(1).strip()
            nombre = limpiar_nombre(match.group(2))
            metadatos = [ws.cell(fila + 1, columna + offset).value for offset in range(4)]
            categoria = CODIGOS_CATEGORIA.get(str(metadatos[0]).strip(), "SIN_CATEGORIA")

            creditos = 0
            for item in metadatos:
                if isinstance(item, (int, float)):
                    creditos = int(item)
                    break

            bloque = ""
            if metadatos[1] is not None:
                bloque = str(metadatos[1]).strip()

            mapa[clave] = {
                "clave": clave,
                "nombre": nombre,
                "ciclo": semestre,
                "ciclo_anual": ciclo_anual,
                "categoria": categoria,
                "creditos": creditos,
                "bloque": bloque,
                "requisitos": [],
                "origen": "excel_oficial",
            }

    return dict(sorted(mapa.items()))


def construir_equivalencias_legacy(
    mapa_legacy: Dict[str, Dict],
    mapa_real: Dict[str, Dict],
) -> Tuple[Dict[str, str], Dict[str, List[str]]]:
    equivalencias = {}
    razones = {}

    nombres_reales = {}
    for clave, datos in mapa_real.items():
        nombre_norm = normalizar(datos.get("nombre", ""))
        nombres_reales.setdefault(nombre_norm, []).append(clave)

    for legacy_clave, legacy_datos in mapa_legacy.items():
        legacy_nombre = legacy_datos.get("nombre", "")
        coincidencias = nombres_reales.get(normalizar(legacy_nombre), [])
        if len(coincidencias) == 1:
            equivalencias[legacy_clave] = coincidencias[0]
            razones[legacy_clave] = ["match_nombre_exacto"]

    for legacy_clave, real_clave in EQUIVALENCIAS_MANUALES.items():
        if real_clave in mapa_real:
            equivalencias[legacy_clave] = real_clave
            razones.setdefault(legacy_clave, []).append("equivalencia_manual")

    return equivalencias, razones


def generar_mapa_completo(
    mapa_real: Dict[str, Dict],
    mapa_legacy: Dict[str, Dict],
    equivalencias: Dict[str, str],
    razones_equivalencia: Dict[str, List[str]],
) -> Dict[str, Dict]:
    legacy_a_real = equivalencias
    real_a_legacy = {}
    for legacy_clave, real_clave in legacy_a_real.items():
        real_a_legacy.setdefault(real_clave, []).append(legacy_clave)

    mapa_completo = {}
    for real_clave, datos in mapa_real.items():
        materia = dict(datos)
        legacy_claves = sorted(real_a_legacy.get(real_clave, []))
        requisitos = set()

        for legacy_clave in legacy_claves:
            legacy_datos = mapa_legacy.get(legacy_clave, {})
            for requisito_legacy in legacy_datos.get("requisitos", []):
                requisito_real = legacy_a_real.get(requisito_legacy)
                if requisito_real:
                    requisitos.add(requisito_real)

        materia["requisitos"] = sorted(requisitos)
        materia["legacy_claves"] = legacy_claves
        materia["equivalencia_origen"] = sorted({
            razon
            for legacy_clave in legacy_claves
            for razon in razones_equivalencia.get(legacy_clave, [])
        })

        mapa_completo[real_clave] = materia

    return dict(sorted(mapa_completo.items()))


def guardar_json(ruta: Path, data: Dict) -> None:
    with open(ruta, "w", encoding="utf-8") as archivo:
        json.dump(data, archivo, ensure_ascii=False, indent=2)


def main() -> None:
    mapa_real = leer_mapa_excel(RUTA_EXCEL)

    with open(RUTA_MAPA_LEGACY, "r", encoding="utf-8") as archivo:
        mapa_legacy = json.load(archivo)

    equivalencias, razones_equivalencia = construir_equivalencias_legacy(
        mapa_legacy,
        mapa_real,
    )
    mapa_completo = generar_mapa_completo(
        mapa_real,
        mapa_legacy,
        equivalencias,
        razones_equivalencia,
    )

    equivalencias_detalladas = {}
    for legacy_clave, real_clave in sorted(equivalencias.items()):
        equivalencias_detalladas[legacy_clave] = {
            "real_clave": real_clave,
            "real_nombre": mapa_real[real_clave]["nombre"],
            "razones": razones_equivalencia.get(legacy_clave, []),
        }

    guardar_json(RUTA_SALIDA_BASE, mapa_real)
    guardar_json(RUTA_SALIDA_COMPLETA, mapa_completo)
    guardar_json(RUTA_EQUIVALENCIAS, equivalencias_detalladas)

    total_semestres = {}
    for datos in mapa_real.values():
        semestre = datos["ciclo"]
        total_semestres[semestre] = total_semestres.get(semestre, 0) + 1

    print(f"Mapa real generado: {len(mapa_real)} materias")
    print(f"Mapa enriquecido generado: {len(mapa_completo)} materias")
    print(f"Equivalencias legacy->real: {len(equivalencias_detalladas)}")
    for semestre in sorted(total_semestres):
        print(f"  Semestre {semestre}: {total_semestres[semestre]} materias")
    print(f"Archivo base: {RUTA_SALIDA_BASE}")
    print(f"Archivo completo: {RUTA_SALIDA_COMPLETA}")
    print(f"Equivalencias: {RUTA_EQUIVALENCIAS}")


if __name__ == "__main__":
    main()